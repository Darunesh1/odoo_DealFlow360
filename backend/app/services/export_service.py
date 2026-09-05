"""PDF and XLS exports for the reporting screen (spec A7).

Both are built in memory and streamed; nothing is written to disk, so there is
no temp file to clean up and no path for one request to read another's export.
"""

from datetime import date
from io import BytesIO
from typing import Any, Sequence


def to_xlsx(rows: Sequence[dict[str, Any]], *, title: str = "Sales") -> bytes:
    """One sheet, a header row, and the data. Column widths from the content
    so the file opens readable rather than as a wall of ###."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = title[:31]  # Excel's limit, silently truncated otherwise.

    if not rows:
        sheet["A1"] = "No sales in that period."
        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    headers = list(rows[0].keys())
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        sheet.append([row.get(header) for header in headers])

    for index, header in enumerate(headers, start=1):
        widest = max(
            [len(str(header))] + [len(str(row.get(header, ""))) for row in rows]
        )
        sheet.column_dimensions[get_column_letter(index)].width = min(widest + 3, 40)

    sheet.freeze_panes = "A2"

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def to_pdf(
    summary: dict[str, Any], rows: Sequence[dict[str, Any]], *, title: str = "Sales report"
) -> bytes:
    """A one-page summary followed by the detail, landscape so the table fits."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    buffer = BytesIO()
    document = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
        title=title,
    )
    styles = getSampleStyleSheet()
    story: list[Any] = [
        Paragraph(title, styles["Title"]),
        Paragraph(
            f"Generated {date.today().isoformat()}", styles["Normal"]
        ),
        Spacer(1, 8 * mm),
    ]

    figures = [
        ("Quotes created", summary.get("quotes_created")),
        ("Confirmed", summary.get("quotes_confirmed")),
        ("Conversion", f"{summary.get('conversion_rate', 0)}%"),
        ("Revenue", f"{summary.get('revenue', 0):,.2f}"),
        ("Margin", f"{summary.get('margin', 0):,.2f}"),
        ("Avg discount", f"{summary.get('average_discount', 0)}%"),
        (
            "Avg approval",
            f"{summary.get('average_approval_hours')} h"
            if summary.get("average_approval_hours") is not None
            else "—",
        ),
    ]
    summary_table = Table(
        [[label for label, _ in figures], [str(value) for _, value in figures]],
        hAlign="LEFT",
    )
    summary_table.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#6B7280")),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ("LINEBELOW", (0, 0), (-1, 0), 0.4, colors.HexColor("#E5E7EB")),
            ]
        )
    )
    story.extend([summary_table, Spacer(1, 8 * mm)])

    if rows:
        headers = list(rows[0].keys())
        data = [headers] + [
            [str(row.get(header, "")) for header in headers] for row in rows[:400]
        ]
        detail = Table(data, repeatRows=1, hAlign="LEFT")
        detail.setStyle(
            TableStyle(
                [
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 7),
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F3F4F6")),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#E5E7EB")),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    (
                        "ROWBACKGROUNDS",
                        (0, 1),
                        (-1, -1),
                        [colors.white, colors.HexColor("#FAFAFA")],
                    ),
                ]
            )
        )
        story.append(detail)
        if len(rows) > 400:
            story.extend(
                [
                    Spacer(1, 4 * mm),
                    Paragraph(
                        f"Showing the first 400 of {len(rows)} lines. "
                        "Export to XLSX for the full set.",
                        styles["Italic"],
                    ),
                ]
            )
    else:
        story.append(Paragraph("No sales in that period.", styles["Normal"]))

    document.build(story)
    return buffer.getvalue()
